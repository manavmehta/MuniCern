from openpyxl import Workbook
from openpyxl import load_workbook
#from openpyxl.utils import get_column_letter
from flask import Flask, request, url_for
from flask_bootstrap import Bootstrap

app = Flask(__name__)
Bootstrap(app)
'''
@app.route('/')
def index():
	return render_template('newissue.html')
'''

@app.route("/matching", methods=["POST"])
def matching():
	
	newissue = str(request.form["issue"]).lower()
	#newissue=input().lower()
	newissue=set(newissue.split(' '))

	wb1 = load_workbook('/Users/manavmehta/Desktop/HACK/sheet.xlsx')
	ws = wb1["Sheet1"]

	issues=[]
	for row in ws.iter_rows():
		issues.append(row[2].value.lower().split(' '))

	best=0
	count=0
	for i in range(len(issues)):
		issue=set(issues[i])
		#print(len(newissue.intersection(issue)))
		cnt=len(newissue.intersection(issue))
		if cnt>count:
			count=cnt
			best=i
	return(str(issues[best]))
	
	print("Works")

if __name__=='__main__':
	app.run(debug=False)